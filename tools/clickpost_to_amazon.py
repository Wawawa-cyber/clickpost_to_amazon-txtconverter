import csv
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

OUTPUT_DIR = Path("output")


def ensure_output_dir() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def load_json(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def load_jsonc(path: Path) -> Dict[str, Any]:
    text = path.read_text(encoding="utf-8")
    # Strip // line comments and /* */ block comments
    text = re.sub(r"/\*.*?\*/", "", text, flags=re.S)
    text = re.sub(r"(^|\s)//.*", "", text)
    return json.loads(text)


def merge_settings(mapping_cfg: Dict[str, Any], tool_cfg: Dict[str, Any]) -> Dict[str, Any]:
    # mapping.json may contain only "mappings". The general defaults can live in tools/config.json.
    def first(*vals):
        for v in vals:
            if v is not None:
                return v
        return None

    combined: Dict[str, Any] = {}
    combined["encoding"] = first(mapping_cfg.get("encoding"), tool_cfg.get("encoding"), "auto")
    combined["input_date_format"] = first(mapping_cfg.get("input_date_format"), tool_cfg.get("input_date_format"), "yyyyMMdd")
    combined["output_date_format"] = first(mapping_cfg.get("output_date_format"), tool_cfg.get("output_date_format"), "yyyy-MM-dd")
    combined["sheet"] = first(mapping_cfg.get("sheet"), tool_cfg.get("sheet"))
    combined["header_row_index"] = first(mapping_cfg.get("header_row_index"), tool_cfg.get("header_row_index"))
    combined["start_data_row_index"] = first(mapping_cfg.get("start_data_row_index"), tool_cfg.get("start_data_row_index"))
    combined["header_search_rows"] = first(mapping_cfg.get("header_search_rows"), tool_cfg.get("header_search_rows"), 20)

    # Accept either top-level mappings or nested under "mappings"
    mappings = mapping_cfg.get("mappings")
    if not mappings:
        # If keys look like amazon columns, assume mapping_cfg itself is mappings
        mappings = mapping_cfg
    combined["mappings"] = mappings or {}
    return combined


def parse_date(value: str, in_fmt: str, out_fmt: str) -> str:
    token_map = {
        "yyyy": "%Y",
        "MM": "%m",
        "dd": "%d",
        "HH": "%H",
        "mm": "%M",
        "ss": "%S",
    }
    def to_py(fmt: str) -> str:
        py = fmt
        for k, v in token_map.items():
            py = py.replace(k, v)
        return py

    src = to_py(in_fmt)
    dst = to_py(out_fmt)
    dt = datetime.strptime(value.strip(), src)
    return dt.strftime(dst)


def detect_encoding(path: Path) -> str:
    """Detect file encoding by trying common encodings"""
    encodings = ['utf-8', 'cp932', 'shift_jis', 'utf-8-sig']
    
    for encoding in encodings:
        try:
            with path.open('r', encoding=encoding) as f:
                f.read(1024)  # Try to read first 1024 characters
                return encoding
        except UnicodeDecodeError:
            continue
    
    # If all fail, default to cp932 for Japanese files
    return 'cp932'


def read_input_csv(path: Path, encoding: str) -> List[Dict[str, Any]]:
    # If encoding detection is requested or fails, try to detect
    if encoding == 'auto' or not encoding:
        encoding = detect_encoding(path)
        print(f"検出されたエンコーディング {encoding}")
    
    try:
        with path.open("r", encoding=encoding, newline="") as f:
            reader = csv.reader(f)
            try:
                header = next(reader)
            except StopIteration:
                return []
            rows: List[Dict[str, Any]] = []
            for cells in reader:
                # Normalize row length to header length
                if len(cells) < len(header):
                    cells = cells + [""] * (len(header) - len(cells))
                row = {header[i]: cells[i] for i in range(len(header))}
                row["__cells__"] = list(cells)
                rows.append(row)
            return rows
    except UnicodeDecodeError as e:
        # If specified encoding fails, try auto-detection
        print(f"指定されたエンコーディング'{encoding}' で読み込みに失敗しました。自動検出を試行します..")
        detected_encoding = detect_encoding(path)
        if detected_encoding != encoding:
            print(f"自動検出されたエンコーディング {detected_encoding}")
            return read_input_csv(path, detected_encoding)
        else:
            raise e


def find_header_row(ws, expected_key: str, search_rows: int) -> Optional[int]:
    for r in range(1, search_rows + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if isinstance(val, str) and val.strip() == expected_key:
                return r
    return None


def build_template_col_index(ws, header_row_idx: int) -> Dict[str, int]:
    out: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        key = ws.cell(row=header_row_idx, column=c).value
        if isinstance(key, str) and key.strip():
            out[key.strip()] = c
    return out


def transform_row(src_row: Dict[str, str], mapcfg: Dict[str, Any]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    in_date_fmt = mapcfg.get("input_date_format", "yyyyMMdd")
    out_date_fmt = mapcfg.get("output_date_format", "yyyy-MM-dd")
    mappings: Dict[str, Any] = mapcfg["mappings"]

    def extract_value(raw: Any, mode: str) -> str:
        s = str(raw or "").strip()
        if not s:
            return ""
        if mode in ("paren_inner", "paren", "inside_paren"):
            # Robust across ASCII () and full-width （）
            m = re.search(r"[\(（]([^（）\)]+)[\)）]", s)
            return m.group(1).strip() if m else ""
        if mode in ("paren_before", "before_paren"):
            # Split on first ASCII or full-width open paren
            for ch in ("(", "（"):
                if ch in s:
                    return s.split(ch, 1)[0].strip()
            return s
        if mode in ("after_multiply", "after_times", "qty_after_times"):
            # Find the last multiply-like mark and take digits after it
            marks = ['×', 'x', 'X', 'ｘ', 'Ｘ', '*', '＊']
            idx = -1
            for mmark in marks:
                idx = max(idx, s.rfind(mmark))
            if idx >= 0 and idx + 1 < len(s):
                tail = s[idx + 1 :]
                m2 = re.search(r"(\d+)", tail)
                return m2.group(1) if m2 else tail.strip()
            # Fallback: any trailing number in the string
            m3 = re.search(r"(\d+)$", s)
            return m3.group(1) if m3 else ""
        return s

    for amazon_col, rule in mappings.items():
        if isinstance(rule, dict):
            if "value" in rule:
                out[amazon_col] = rule.get("value", "")
            elif "source" in rule or "index" in rule:
                if "index" in rule:
                    cells = src_row.get("__cells__", [])
                    idx = int(rule["index"])
                    raw = cells[idx] if isinstance(cells, list) and 0 <= idx < len(cells) else ""
                else:
                    raw = src_row.get(rule["source"], "")
                # Optional extraction step (e.g., paren or quantity parsing)
                if isinstance(rule.get("extract"), str):
                    raw = extract_value(raw, rule["extract"])  # type: ignore
                if rule.get("parse_date") and raw:
                    out[amazon_col] = parse_date(str(raw), in_date_fmt, out_date_fmt)
                else:
                    out[amazon_col] = raw
            else:
                out[amazon_col] = ""
        else:
            out[amazon_col] = src_row.get(str(rule), "")
    return out


def write_to_template(template_path: Path, output_path: Path, records: List[Dict[str, Any]], mapcfg: Dict[str, Any]) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(template_path)
    ws = wb[mapcfg["sheet"]] if mapcfg.get("sheet") else wb.active

    header_row_idx: Optional[int] = mapcfg.get("header_row_index")
    if not header_row_idx:
        header_row_idx = find_header_row(ws, "order-id", mapcfg.get("header_search_rows", 20))
        if not header_row_idx:
            raise RuntimeError("Could not locate header row containing 'order-id'. Specify header_row_index in config/mapping.")

    template_cols = build_template_col_index(ws, header_row_idx)
    missing_cols = [k for k in mapcfg["mappings"].keys() if k not in template_cols]
    if missing_cols:
        raise RuntimeError("Template is missing required columns: " + ", ".join(missing_cols))

    start_row = mapcfg.get("start_data_row_index", header_row_idx + 1)
    row_idx = start_row
    for rec in records:
        transformed = transform_row(rec, mapcfg)
        for col_name, value in transformed.items():
            col = template_cols[col_name]
            ws.cell(row=row_idx, column=col, value=value)
        row_idx += 1

    if output_path.resolve() == template_path.resolve():
        raise RuntimeError("Refusing to overwrite template. Use a different output path.")
    wb.save(output_path)


def to_rows_in_order(records: List[Dict[str, Any]], columns: List[str]) -> List[List[Any]]:
    out: List[List[Any]] = []
    out.append(columns)
    for rec in records:
        out.append([rec.get(col, "") for col in columns])
    return out


def write_csv_utf8(path: Path, rows: List[List[Any]], delimiter: str = ",") -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter=delimiter)
        for row in rows:
            writer.writerow(row)


def write_amazon_format_txt(path: Path, rows: List[List[Any]]) -> None:
    """Write Amazon shipping confirmation format with Shift_JIS encoding and fixed-width tabs"""
    # Amazon format requires exactly 256 tab-separated columns (255 tabs)
    TOTAL_COLUMNS = 256
    
    with path.open("w", encoding="cp932", newline="") as f:
        for i, row in enumerate(rows):
            # Convert all values to string
            line_data = [str(cell) for cell in row]
            
            # Pad with empty strings to reach exactly TOTAL_COLUMNS
            while len(line_data) < TOTAL_COLUMNS:
                line_data.append("")
            
            # Join with tabs and write with CRLF
            line = "\t".join(line_data)
            f.write(line + "\r\n")
        # Add a final empty line to match sample/test Flat.File.ShippingConfirmation.jp.txt
        f.write("\r\n")


def write_tab_separated_txt_with_spacing(path: Path, rows: List[List[Any]]) -> None:
    """Write tab-separated text file with 14 line breaks between each data row"""
    with path.open("w", encoding="utf-8", newline="") as f:
        for i, row in enumerate(rows):
            # Convert all values to string and join with tabs
            line = "\t".join(str(cell) for cell in row)
            f.write(line + "\n")
            
            # Add 14 empty lines between rows (except for the last row)
            if i < len(rows) - 1:
                for _ in range(14):
                    f.write("\n")


def main():
    import argparse
    
    try:
        parser = argparse.ArgumentParser(description="Convert CSV files to Amazon Shipping Confirmation format (Shift_JIS, tab-separated, 237 columns)")
        parser.add_argument("--input", default=None, help="Input CSV file path (if not specified, auto-detects from input/ folder)")
        parser.add_argument("--columns", default=None, help="Template columns JSON path (UTF-8)")
        parser.add_argument("--mapping", default="config/mapping.json", help="Mapping JSON path (UTF-8)")
        parser.add_argument("--config", default="config/config.json", help="Tool config (JSON)")
        args = parser.parse_args()
        
        print("=== CSV → Amazon 出荷通知 変換ツール ===")
        print(f"マッピング設定: {args.mapping}")
        print("変換処理を開始します..")
        print()
        
    except Exception as e:
        print(f"⚠ エラー: 引数の解析に失敗しました - {e}")
        exit(1)

    try:
        # Load configs
        default_tool_cfg: Dict[str, Any] = {
            "output": {
                "write_xlsx": False,
                "write_csv": False,
                "write_txt": True,
                "csv_name": "shipping_confirmation.csv",
                "txt_name": "shipping_confirmation.txt",
                "xlsx_name": "AmazonShippingConfirmation_output.xlsx",
            },
            "io": {
                "encoding": "auto",
                "columns_path": "config/template_columns.json",
            },
            "format": {
                "input_date_format": "yyyyMMdd",
                "output_date_format": "yyyy-MM-dd",
            }
        }

        # Load tool config; prefer JSON, then JSONC fallback, then legacy tools/config.json
        cfg_path = Path(args.config)
        tool_cfg_raw: Dict[str, Any] = {}
        if cfg_path.exists():
            tool_cfg_raw = load_json(cfg_path)
        else:
            alt = Path("config/config.jsonc")
            if alt.exists():
                tool_cfg_raw = load_jsonc(alt)
            else:
                legacy = Path("tools/config.json")
                if legacy.exists():
                    tool_cfg_raw = load_json(legacy)

        # Merge default -> tool_cfg_raw
        def deep_merge(a: Dict[str, Any], b: Dict[str, Any]) -> Dict[str, Any]:
            out = dict(a)
            for k, v in b.items():
                if isinstance(v, dict) and isinstance(out.get(k), dict):
                    out[k] = deep_merge(out[k], v)  # type: ignore
                else:
                    out[k] = v
            return out

        merged_tool_cfg = deep_merge(default_tool_cfg, tool_cfg_raw)

        mapping_cfg_raw = {}
        mapping_path = Path(args.mapping)
        if mapping_path.exists():
            mapping_cfg_raw = load_json(mapping_path)

        # Flatten merged tool cfg for merge_settings compatibility
        flat_tool_cfg: Dict[str, Any] = {}
        flat_tool_cfg.update(merged_tool_cfg.get("io", {}))
        flat_tool_cfg.update(merged_tool_cfg.get("format", {}))
        flat_tool_cfg.update(merged_tool_cfg.get("output", {}))
        cfg = merge_settings(mapping_cfg_raw, flat_tool_cfg)

        # Ensure standard folders
        Path("input").mkdir(parents=True, exist_ok=True)
        ensure_output_dir()

        # Auto-detect input CSV file if not specified
        input_path = args.input
        if not input_path:
            input_dir = Path("input")
            csv_files = list(input_dir.glob("*.csv"))
            if not csv_files:
                raise FileNotFoundError("input フォルダにCSVファイルが見つかりません")
            input_path = str(csv_files[0])
            print(f"自動検出されたファイル: {input_path}")
        
        # Check if input file exists
        input_file_path = Path(input_path)
        if not input_file_path.exists():
            raise FileNotFoundError(f"指定されたファイルが見つかりません: {input_path}")
        
        print(f"入力ファイル: {input_path}")

        # Read input CSV
        rows = read_input_csv(input_file_path, cfg.get("encoding", "auto"))

        # Load template columns order from JSON (Excel template no longer used)
        columns_path = Path(args.columns or merged_tool_cfg.get("io", {}).get("columns_path", "config/template_columns.json"))
        if not columns_path.exists():
            raise SystemExit(f"Columns JSON not found: {columns_path}")
        cols_json = load_json(columns_path)
        if isinstance(cols_json, dict):
            header_rows = cols_json.get("header_rows", [])
            columns = cols_json.get("columns", [])
        elif isinstance(cols_json, list):
            header_rows = [cols_json]
            columns = cols_json
        else:
            raise SystemExit("Invalid columns JSON format. Expect list or {header_rows, columns} object.")

        # Transform and prepare output
        transformed: List[Dict[str, Any]] = [transform_row(r, cfg) for r in rows]
        # Build matrix with header rows first (if provided), then header line (columns), then data
        matrix: List[List[Any]] = []
        for hr in header_rows:
            # hr may be shorter; write as-is
            matrix.append(list(hr))
        matrix.append(columns)
        for rec in transformed:
            matrix.append([rec.get(col, "") for col in columns])

        # output dir already ensured

        # Generate timestamped filenames
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Initialize file paths
        txt_file_path = None
        csv_file_path = None
        
        # Write TXT (tab) in UTF-8
        if merged_tool_cfg.get("output", {}).get("write_txt", True):
            base_name = merged_tool_cfg.get("output", {}).get("txt_name", "shipping_confirmation.txt")
            name_parts = base_name.rsplit('.', 1)
            if len(name_parts) == 2:
                timestamped_name = f"{name_parts[0]}_{timestamp}.{name_parts[1]}"
            else:
                timestamped_name = f"{base_name}_{timestamp}"
            txt_file_path = OUTPUT_DIR / timestamped_name
            write_amazon_format_txt(txt_file_path, matrix)

        # Optionally write CSV
        if merged_tool_cfg.get("output", {}).get("write_csv", False):
            base_name = merged_tool_cfg.get("output", {}).get("csv_name", "shipping_confirmation.csv")
            name_parts = base_name.rsplit('.', 1)
            if len(name_parts) == 2:
                timestamped_name = f"{name_parts[0]}_{timestamp}.{name_parts[1]}"
            else:
                timestamped_name = f"{base_name}_{timestamp}"
            csv_file_path = OUTPUT_DIR / timestamped_name
            write_csv_utf8(csv_file_path, matrix, ",")

        # XLSX output deprecated (Excel template not used anymore)
        if merged_tool_cfg.get("output", {}).get("write_xlsx", False):
            print("[warn] write_xlsx is no longer supported because Excel template is not used. Skipped.")

        print("Done. Outputs in:", OUTPUT_DIR)
        
        # 成功メッセージの表示
        if merged_tool_cfg.get("output", {}).get("write_txt", True) and txt_file_path:
            if txt_file_path.exists():
                print(f"[成功] 出力完了: {txt_file_path}")
                print(f"[成功] 処理済み行数: {len(transformed)}")
            else:
                print("[エラー] 出力ファイルが見つかりません")
                exit(1)
        
        if merged_tool_cfg.get("output", {}).get("write_csv", False) and csv_file_path:
            if csv_file_path.exists():
                print(f"[成功] CSV出力完了: {csv_file_path}")
        
    except FileNotFoundError as e:
        print(f"[エラー] ファイルが見つかりません - {e}")
        print("確認してください:")
        print(f"  - {args.input} (入力ファイル)")
        print(f"  - {args.mapping} (マッピング設定)")
        print(f"  - {args.config} (設定ファイル)")
        exit(1)
    except json.JSONDecodeError as e:
        print(f"[エラー] JSON設定ファイルの形式が正しくありません - {e}")
        exit(1)
    except UnicodeDecodeError as e:
        print(f"[エラー] ファイルエンコーディングが正しくありません - {e}")
        print("入力ファイルのエンコード形式を確認してください（UTF-8、Shift_JIS、CP932など）")
        exit(1)
    except Exception as e:
        print(f"[エラー] 予期しないエラーが発生しました - {e}")
        exit(1)


if __name__ == "__main__":
    main()
